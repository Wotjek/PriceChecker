#!/usr/bin/env python3
"""
Price Tracker - monitoring najnizszych cen w sklepach Europy kontynentalnej.

Pipeline:
  1. DISCOVERY  - Google Programmable Search (EAN / frazy) -> nowe URL-e sklepow
  2. MONITORING - pobranie stron, ekstrakcja ceny+dostepnosci z JSON-LD / meta
  3. FX         - przeliczenie na PLN po kursie NBP z danego dnia (fallback: frankfurter.app)
  4. HISTORIA   - dopisanie wiersza (data, produkt, najnizsza cena, sklep, URL) do CSV
  5. RAPORT     - regeneracja pliku Excel z historia i wykresem dla kazdego produktu
"""

import csv
import json
import os
import re
import statistics
import sys
import time
from datetime import date
from pathlib import Path
from urllib.parse import parse_qsl, urlencode, urljoin, urlparse, urlunparse

import requests
import yaml
from bs4 import BeautifulSoup

ROOT = Path(__file__).parent
DATA = ROOT / "data"
OUTPUT = ROOT / "output"
PRODUCTS_FILE = ROOT / "products.yaml"
SOURCES_FILE = DATA / "sources.json"      # baza znanych URL-i per produkt
HISTORY_FILE = DATA / "history.csv"       # najlepsza cena / produkt / dzien
OFFERS_FILE = DATA / "all_offers.csv"     # wszystkie znalezione oferty (audyt)
QUOTA_FILE = DATA / "serpapi_quota.json"  # stan limitu darmowych zapytan SerpAPI
SERPER_QUOTA_FILE = DATA / "serper_quota.json"  # lokalny licznik kredytow Serper
BLIND_FILE = DATA / "blind_spots.json"  # sklepy bez ceny mimo Google (per run)
EXCEL_FILE = OUTPUT / "prices.xlsx"

# Naglowki jak w prawdziwym Chrome - czesc sklepow (bike24, r2-bike, allegro)
# odrzuca requesty z golym User-Agentem. Accept-Encoding ustawia requests
# samodzielnie (z pakietem Brotli takze "br", jak przegladarka).
# Kolejnosc jezykow jak dotychczas (en > de > pl), zeby sklepy wielojezyczne
# nie zaczely nagle serwowac innych wersji/walut niz w historii.
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/143.0.0.0 Safari/537.36")
BROWSER_HEADERS = {
    "User-Agent": UA,
    "Accept": ("text/html,application/xhtml+xml,application/xml;q=0.9,"
               "image/avif,image/webp,image/apng,*/*;q=0.8,"
               "application/signed-exchange;v=b3;q=0.7"),
    "Accept-Language": "en-US,en;q=0.9,de;q=0.8,pl;q=0.7",
    "sec-ch-ua": '"Not)A;Brand";v="8", "Chromium";v="143", "Google Chrome";v="143"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"Windows"',
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
    "Upgrade-Insecure-Requests": "1",
}
TIMEOUT = 25
BROWSER_TIMEOUT = 40   # headless Chromium: strona + ewentualny challenge JS
SERPAPI_TIMEOUT = 60   # SerpAPI (zwlaszcza Shopping) bywa wolne; timeout i tak zuzywa limit
SLEEP_BETWEEN_FETCHES = 2.0

# --- Europa kontynentalna: dozwolone waluty i TLD ---------------------------
ALLOWED_CURRENCIES = {"PLN", "EUR", "CZK", "DKK", "SEK", "NOK", "CHF",
                      "HUF", "RON", "BGN"}
EU_TLDS = {"pl", "de", "nl", "fr", "it", "es", "be", "at", "cz", "sk", "dk",
           "se", "fi", "no", "pt", "hu", "ro", "si", "hr", "lt", "lv", "ee",
           "gr", "bg", "lu", "ch", "li", "ie", "eu"}
NEUTRAL_TLDS = {"com", "net", "org", "shop", "bike", "cc", "io", "store"}
BLOCKED_TLDS = {"uk", "gg", "je", "im", "us", "ca", "au", "nz", "jp", "cn",
                "in", "br", "mx", "za", "kr", "sg", "hk", "tr", "ru", "by"}
BLOCKED_DOMAINS = {"google.com", "youtube.com",
                   "facebook.com", "reddit.com", "pinterest.com",
                   # marketplace'y - ceny aukcyjne/uzywane, nie sklepowe
                   "ebay.com", "ebay.de", "ebay.pl", "ebay.it", "ebay.fr",
                   "ebay.es", "ebay.nl", "ebay.at", "ebay.ch",
                   "amazon.com", "amazon.de", "amazon.it", "amazon.fr",
                   "amazon.es", "amazon.pl", "amazon.nl", "aliexpress.com",
                   # porownywarki cen - agregatory, nie sklepy (ceny bywaja stare)
                   "idealo.de", "idealo.pl", "idealo.it", "ceneo.pl",
                   "arukereso.hu", "geizhals.de", "geizhals.at", "geizhals.eu",
                   "guenstiger.de", "billiger.de", "trovaprezzi.it",
                   "skroutz.gr", "pricerunner.com", "pricerunner.dk",
                   "pricerunner.se",
                   # strony producentow bez cen sklepowych
                   "salsacycles.com", "dtswiss.com", "sram.com",
                   # portale nieruchomosci (kolizje nazw typu "FM2797" = droga w Teksasie)
                   "homes.com", "redfin.com", "har.com", "compass.com",
                   "zillow.com", "realtor.com", "trulia.com",
                   # media/recenzje/fora - artykuly, nie sklepy
                   "fat-bike.com", "singletracks.com", "bikeradar.com",
                   "pinkbike.com", "bikerumor.com", "cyclingnews.com",
                   "mtb-news.de", "99spokes.com", "scribd.com",
                   "mtbr.com", "expeditionportal.com",
                   # social / katalogi / rejestry - nie sklepy
                   "instagram.com", "yelp.com", "upcindex.com", "rejestr.io",
                   # marketplace'y aukcyjne i ogloszenia
                   "allegro.pl", "allegro.hu", "allegro.cz", "allegro.sk",
                   "fillaritori.com",
                   # sklepy nie-rowerowe (kolizje EAN/kodow produktow)
                   "memorialtacostx.com", "americanpharmawholesale.com",
                   "bbqguys.com", "staples.com", "bestbuy.com",
                   "dkhardware.com"}
BLOCKED_PATH_RE = re.compile(r"/(blogs?|news|review|artykul|magazin)e?s?/", re.I)

IN_STOCK = {"instock", "limitedavailability", "onlineonly"}


# ============================================================ helpers =======

LOG_BUFFER = []

# maskowanie sekretow w logach: log trafia do repo i do dashboardu
# publikowanego na (publicznych) GitHub Pages - klucz API nie moze wyciec,
# np. w komunikacie bledu HTTP z pelnym URL-em zapytania
_SECRET_PARAM_RE = re.compile(r"(api_key|apikey|key|token)=[^&\s\"']+", re.I)
_SECRET_ENV_VARS = ("SERPAPI_KEY", "TAVILY_API_KEY", "BRAVE_API_KEY",
                    "GOOGLE_API_KEY")


def _mask_secrets(s):
    s = _SECRET_PARAM_RE.sub(r"\1=***", s)
    for var in _SECRET_ENV_VARS:
        v = os.environ.get(var, "").strip()
        if v:
            s = s.replace(v, "***")
    return s


def log(msg):
    msg = _mask_secrets(str(msg))
    print(msg, flush=True)
    LOG_BUFFER.append(msg)


def save_run_log():
    DATA.mkdir(exist_ok=True)
    (DATA / "last_run.log").write_text("\n".join(LOG_BUFFER) + "\n",
                                       encoding="utf-8")


def domain_of(url):
    d = urlparse(url).netloc.lower()
    return d[4:] if d.startswith("www.") else d


# parametry sledzace w URL-ach (Google/e-mail/ads) - bez wplywu na tresc strony;
# ich usuniecie scala duplikaty typu ?srsltid=... w bazie zrodel i audycie
TRACKING_PARAM_RE = re.compile(
    r"^(utm_|mc_|pk_|piwik_)|^(srsltid|gclid|dclid|fbclid|msclkid|igshid|"
    r"ttclid|yclid|wbraid|gbraid|_ga|_gl|mkt_tok|sc_src|sc_uid)$", re.I)


def normalize_url(url):
    """Usuwa parametry sledzace i fragment (#...) - kanoniczna postac URL-a."""
    try:
        parts = urlparse(url)
        q = [(k, v) for k, v in parse_qsl(parts.query, keep_blank_values=True)
             if not TRACKING_PARAM_RE.match(k)]
        return urlunparse(parts._replace(query=urlencode(q), fragment=""))
    except Exception:
        return url


def normalize_sources(sources):
    """Migracja bazy URL-i: scala wpisy rozniace sie tylko parametrami sledzacymi."""
    merged = 0
    for pid, urls in sources.items():
        clean = {}
        for url, meta in urls.items():
            nu = normalize_url(url)
            if nu in clean:
                merged += 1
            else:
                clean[nu] = meta
        sources[pid] = clean
    if merged:
        log(f"[NORMALIZACJA] Scalono {merged} zduplikowanych URL-i "
            f"(parametry sledzace)")
    return sources


def tld_of(domain):
    return domain.rsplit(".", 1)[-1]


WORLDWIDE_CURRENCIES = ALLOWED_CURRENCIES | {"USD", "GBP", "CAD", "AUD"}


def url_allowed(url, extra_excluded, worldwide=False):
    if BLOCKED_PATH_RE.search(urlparse(url).path or ""):
        return False
    return domain_allowed(domain_of(url), extra_excluded, worldwide)


def domain_allowed(domain, extra_excluded, worldwide=False):
    if not domain:
        return False
    # dopasowanie z subdomenami: kerekagy.arukereso.hu, support.sram.com itd.
    for blocked in (BLOCKED_DOMAINS, extra_excluded):
        if any(domain == b or domain.endswith("." + b) for b in blocked):
            return False
    if worldwide:
        return True  # produkt globalny (np. dostepny glownie w US)
    if any(domain.endswith("." + b) or tld_of(domain) == b for b in BLOCKED_TLDS):
        return False
    tld = tld_of(domain)
    return tld in EU_TLDS or tld in NEUTRAL_TLDS


# ======================================================== FX (kurs dnia) ====

def get_fx_rates():
    """Zwraca slownik {waluta: kurs_PLN} z NBP (tabela A), fallback frankfurter.app."""
    rates = {"PLN": 1.0}
    try:
        r = requests.get("https://api.nbp.pl/api/exchangerates/tables/A?format=json",
                         timeout=TIMEOUT)
        r.raise_for_status()
        for row in r.json()[0]["rates"]:
            rates[row["code"]] = float(row["mid"])
        log(f"[FX] Kursy NBP pobrane ({len(rates)} walut)")
        return rates
    except Exception as e:
        log(f"[FX] NBP niedostepne ({e}), probuje frankfurter.app")
    try:
        r = requests.get("https://api.frankfurter.app/latest?from=PLN", timeout=TIMEOUT)
        r.raise_for_status()
        for code, v in r.json()["rates"].items():
            rates[code] = 1.0 / float(v)  # kurs waluty w PLN
        log(f"[FX] Kursy frankfurter pobrane ({len(rates)} walut)")
    except Exception as e:
        log(f"[FX] BLAD: brak kursow walut ({e}) - tylko oferty w PLN beda uzyte")
    return rates


# ========================================================== discovery =======

# Klucz odrzucony (401/403) powtarza sie dla KAZDEGO zapytania - bez sensu
# mielic setki blednych wywolan i smiecic log. Po pierwszym odrzuceniu silnik
# jest wylaczany do konca runu, z jedna czytelna wskazowka co naprawic.
_auth_dead = set()


def _auth_fail(engine, err, hint):
    """True gdy blad to odrzucony klucz - wylacza silnik do konca runu."""
    code = getattr(getattr(err, "response", None), "status_code", None)
    if code in (401, 402, 403):
        if engine not in _auth_dead:
            _auth_dead.add(engine)
            log(f"[AUTH] {engine}: klucz odrzucony (HTTP {code}) - wylaczam "
                f"silnik do konca runu. {hint}")
        return True
    return False


_HINT_SERPAPI = ("Limit SerpAPI wyczerpany albo zly sekret SERPAPI_KEY "
                 "(serpapi.com -> Api Key).")
_HINT_SERPER = ("Sprawdz sekret SERPER_API_KEY (serper.dev -> API keys); "
                "wartosc bez cudzyslowow i spacji.")
_HINT_GOOGLE = ("Sprawdz GOOGLE_API_KEY: w Google Cloud wlacz Custom Search "
                "API i ustaw Application restrictions klucza na None "
                "(restrykcja 'Websites' odrzuca wywolania z serwera).")


def google_search(query, api_key, cx, pages=1):
    """Google Programmable Search (JSON API) - wyszukiwarka PSE uzytkownika.
    Od 2026-01-20 nowe PSE nie moga przeszukiwac calej sieci - wyniki tylko
    z domen dopisanych do wyszukiwarki (max 50)."""
    urls = []
    if "google" in _auth_dead:
        return urls
    for page in range(pages):
        params = {"key": api_key, "cx": cx, "q": query,
                  "num": 10, "start": 1 + page * 10}
        try:
            r = requests.get("https://www.googleapis.com/customsearch/v1",
                             params=params, timeout=TIMEOUT)
            if r.status_code == 429:
                log("[DISCOVERY][google] Limit dzienny wyczerpany")
                return urls
            r.raise_for_status()
            items = r.json().get("items", [])
            urls += [it["link"] for it in items if "link" in it]
            if len(items) < 10:
                break
        except Exception as e:
            if not _auth_fail("google", e, _HINT_GOOGLE):
                log(f"[DISCOVERY][google] Blad zapytania '{query}': {e}")
            break
        time.sleep(0.5)
    return urls


def serpapi_search(query, api_key, gl=None):
    """SerpAPI - prawdziwe wyniki Google (plan darmowy: 250 zapytan/mies.).
    gl = kod kraju (np. 'de', 'pl') - wyniki z lokalnej wersji Google."""
    if "serpapi" in _auth_dead:
        return []
    try:
        params = {"engine": "google", "q": query, "num": 20, "api_key": api_key}
        if gl:
            params.update({"gl": gl, "google_domain": f"google.{gl}"})
        r = requests.get("https://serpapi.com/search.json",
                         params=params, timeout=SERPAPI_TIMEOUT)
        r.raise_for_status()
        j = r.json()
        if j.get("error"):
            log(f"[DISCOVERY][serpapi] {j['error']}")
            return []
        return [it["link"] for it in j.get("organic_results", []) if it.get("link")]
    except Exception as e:
        if not _auth_fail("serpapi", e, _HINT_SERPAPI):
            log(f"[DISCOVERY][serpapi] Blad zapytania '{query}': {e}")
        return []


def serpapi_shopping_items(query, api_key, gl=None):
    """SerpAPI Google Shopping - surowe oferty (tytul, cena, sklep, link).
    Zrodlem cen jest feed Merchant Center wysylany Google przez same sklepy,
    wiec dziala takze dla sklepow blokujacych boty (bike24, r2-bike)."""
    if "serpapi" in _auth_dead:
        return []
    try:
        # Shopping nie radzi sobie z cudzyslowami (zwraca "no results")
        params = {"engine": "google_shopping", "q": query.replace('"', ""),
                  "api_key": api_key}
        if gl:
            params["gl"] = gl
        r = requests.get("https://serpapi.com/search.json",
                         params=params, timeout=SERPAPI_TIMEOUT)
        r.raise_for_status()
        j = r.json()
        if j.get("error"):
            log(f"[SHOPPING] {j['error']}")
            return []
        return j.get("shopping_results", [])
    except Exception as e:
        if not _auth_fail("serpapi", e, _HINT_SERPAPI):
            log(f"[SHOPPING] Blad zapytania '{query}': {e}")
        return []


def serpapi_shopping_search(query, api_key, gl=None):
    """Discovery przez Shopping: same linki do stron sklepow."""
    out = []
    for it in serpapi_shopping_items(query, api_key, gl):
        link = it.get("link") or ""
        # pomijamy linki do karty produktu Google (nie sklepu)
        if link.startswith("http") and "google." not in domain_of(link):
            out.append(link)
    return out


def tavily_search(query, api_key):
    """Tavily - web search API (plan darmowy: ~1000 zapytan/mies.)."""
    try:
        r = requests.post("https://api.tavily.com/search",
                          json={"api_key": api_key, "query": query,
                                "max_results": 20}, timeout=TIMEOUT)
        r.raise_for_status()
        return [it["url"] for it in r.json().get("results", []) if it.get("url")]
    except Exception as e:
        log(f"[DISCOVERY][tavily] Blad zapytania '{query}': {e}")
        return []


def brave_search(query, api_key):
    """Brave Search API (plan platny/kredytowy)."""
    try:
        r = requests.get("https://api.search.brave.com/res/v1/web/search",
                         params={"q": query, "count": 20},
                         headers={"X-Subscription-Token": api_key,
                                  "Accept": "application/json"}, timeout=TIMEOUT)
        r.raise_for_status()
        return [it["url"] for it in r.json().get("web", {}).get("results", [])
                if it.get("url")]
    except Exception as e:
        log(f"[DISCOVERY][brave] Blad zapytania '{query}': {e}")
        return []


# Serper nie ma endpointu stanu konta, a kredyt schodzi za kazda udana
# odpowiedz - tracker jest jedynym konsumentem klucza, wiec zlicza zuzycie
# sam i utrwala kumulatywnie w data/serper_quota.json (save_serper_quota)
_serper_used = {"n": 0}


def serper_shopping_items(query, api_key, gl=None):
    """Serper.dev Google Shopping - te same dane co serpapi_shopping_items
    (feed Merchant Center), inny dostawca: 2500 darmowych zapytan bez karty.
    Ceny tylko jako tekst ('3.499,00 EUR') - parsowane przy dopasowaniu."""
    if "serper" in _auth_dead:
        return []
    try:
        payload = {"q": query.replace('"', "")}
        if gl:
            payload["gl"] = gl
        r = requests.post("https://google.serper.dev/shopping",
                          headers={"X-API-KEY": api_key,
                                   "Content-Type": "application/json"},
                          json=payload, timeout=TIMEOUT)
        r.raise_for_status()
        _serper_used["n"] += 1
        return r.json().get("shopping", [])
    except Exception as e:
        if not _auth_fail("serper", e, _HINT_SERPER):
            log(f"[SHOPPING][serper] Blad zapytania '{query}': {e}")
        return []


def serper_search(query, api_key, gl=None):
    """Serper.dev - organiczne wyniki Google (discovery)."""
    if "serper" in _auth_dead:
        return []
    try:
        payload = {"q": query, "num": 20}
        if gl:
            payload["gl"] = gl
        r = requests.post("https://google.serper.dev/search",
                          headers={"X-API-KEY": api_key,
                                   "Content-Type": "application/json"},
                          json=payload, timeout=TIMEOUT)
        r.raise_for_status()
        _serper_used["n"] += 1
        return [it["link"] for it in r.json().get("organic", [])
                if it.get("link")]
    except Exception as e:
        if not _auth_fail("serper", e, _HINT_SERPER):
            log(f"[DISCOVERY][serper] Blad zapytania '{query}': {e}")
        return []


def _discovery_engines():
    """Buduje liste dostepnych silnikow na podstawie sekretow w env."""
    engines = []
    if os.environ.get("SERPAPI_KEY", "").strip():
        k = os.environ["SERPAPI_KEY"].strip()
        engines.append(("serpapi", lambda q, gl=None: serpapi_search(q, k, gl)))
        engines.append(("serpapi_shopping",
                        lambda q, gl=None: serpapi_shopping_search(q, k, gl)))
    if os.environ.get("SERPER_API_KEY", "").strip():
        k = os.environ["SERPER_API_KEY"].strip()
        engines.append(("serper", lambda q, gl=None: serper_search(q, k, gl)))
    if os.environ.get("TAVILY_API_KEY", "").strip():
        k = os.environ["TAVILY_API_KEY"].strip()
        engines.append(("tavily", lambda q: tavily_search(q, k)))
    if os.environ.get("BRAVE_API_KEY", "").strip():
        k = os.environ["BRAVE_API_KEY"].strip()
        engines.append(("brave", lambda q: brave_search(q, k)))
    gk, cx = os.environ.get("GOOGLE_API_KEY", "").strip(), os.environ.get("GOOGLE_CX", "").strip()
    if gk and cx:
        engines.append(("google", lambda q: google_search(q, gk, cx)))
    return engines


def run_discovery(products, sources, settings):
    # Light FIRE: tylko monitoring znanych URL-i, zero zapytan do wyszukiwarek
    if os.environ.get("SKIP_DISCOVERY", "").strip().lower() in ("1", "true", "yes"):
        log("[DISCOVERY] Light FIRE - pomijam discovery, tylko monitoring "
            "znanych URL-i (limit SerpAPI nietkniety)")
        return sources
    engines = _discovery_engines()
    if not engines:
        log("[DISCOVERY] Brak kluczy (SERPAPI_KEY / TAVILY_API_KEY / BRAVE_API_KEY / "
            "GOOGLE_API_KEY+CX) - pomijam discovery, uzywam znanych URL-i")
        return sources

    # oszczedzanie limitu: przy trybie weekly harmonogram robi discovery tylko
    # w poniedzialki; reczny FIRE (workflow_dispatch) i run lokalny - zawsze
    mode = str(settings.get("discovery") or "weekly").lower()
    event = os.environ.get("GITHUB_EVENT_NAME", "")
    if event == "schedule" and mode == "weekly" and date.today().weekday() != 0:
        log(f"[DISCOVERY] Tryb weekly - dzis pomijam (silniki: "
            f"{[n for n, _ in engines]}), monitoring dziala normalnie")
        return sources

    log(f"[DISCOVERY] Silniki: {[n for n, _ in engines]}")
    today = date.today().isoformat()
    for p in products:
        pid = p["id"]
        known = sources.setdefault(pid, {})
        excluded = set(p.get("exclude_domains") or [])
        worldwide = bool(p.get("worldwide"))
        queries = []
        if p.get("ean"):
            queries.append(f'"{p["ean"]}"')
        queries += p.get("queries") or []
        if not queries:
            queries = [f'"{p["name"]}"']

        # produkty EU: pytaj lokalne wersje Google (serpapi); worldwide: domyslna (US)
        eu_countries = settings.get("serpapi_countries") or ["de", "pl"]
        # Serper ma osobna, duza pule zapytan - moze przeszukiwac szersza
        # liste krajow Europy bez obciazania limitu SerpAPI
        serper_countries = settings.get("discovery_countries") or eu_countries
        found = 0
        for q in queries:
            for name, fn in engines:
                # oszczedzanie limitu SerpAPI: Shopping tylko dla pierwszego
                # (najprecyzyjniejszego) zapytania - EAN, gdy jest podany
                if name == "serpapi_shopping" and q != queries[0]:
                    continue
                # silniki Google (serpapi*, serper) obsluguja gl= per kraj
                country_list = (serper_countries if name == "serper"
                                else eu_countries)
                countries = ([None] if (worldwide or not name.startswith("serp"))
                             else list(country_list))
                for gl in countries:
                    urls = fn(q, gl=gl) if name.startswith("serp") else fn(q)
                    for url in urls:
                        url = normalize_url(url)
                        dom = domain_of(url)
                        if url not in known and url_allowed(url, excluded, worldwide):
                            known[url] = {"domain": dom, "first_seen": today,
                                          "via": name + (f":{gl}" if gl else "")}
                            found += 1
                    time.sleep(0.4)
        log(f"[DISCOVERY] {pid}: +{found} nowych URL-i (razem {len(known)})")
    return sources


def save_serpapi_quota():
    """Zapisuje stan limitu SerpAPI (endpoint /account nie zuzywa zapytan)."""
    key = os.environ.get("SERPAPI_KEY", "").strip()
    if not key:
        return
    try:
        r = requests.get("https://serpapi.com/account",
                         params={"api_key": key}, timeout=TIMEOUT)
        r.raise_for_status()
        j = r.json()
        data = {"plan": j.get("plan_name", ""),
                "per_month": j.get("searches_per_month"),
                "used": j.get("this_month_usage"),
                "left": j.get("plan_searches_left"),
                "checked": date.today().isoformat()}
        DATA.mkdir(exist_ok=True)
        QUOTA_FILE.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
        log(f"[SERPAPI] Limit: pozostalo {data['left']} z {data['per_month']} "
            f"(zuzyte w tym miesiacu: {data['used']})")
    except Exception as e:
        log(f"[SERPAPI] Nie udalo sie pobrac stanu limitu: {e}")


def save_serper_quota(settings):
    """Utrwala lokalny licznik kredytow Serper (brak API stanu konta).
    Pula startowa/dokupiona: settings.serper_credits (domyslnie 2500);
    po dokupieniu kredytow zaktualizuj settings i usun serper_quota.json."""
    if _serper_used["n"] == 0 and not SERPER_QUOTA_FILE.exists():
        return  # Serper nieuzywany i nie byl uzywany - nic do pokazania
    used = 0
    try:
        used = int(json.loads(SERPER_QUOTA_FILE
                              .read_text(encoding="utf-8")).get("used") or 0)
    except Exception:
        pass
    used += _serper_used["n"]
    total = int(settings.get("serper_credits") or 2500)
    data = {"total": total, "used": used, "left": max(0, total - used),
            "this_run": _serper_used["n"],
            "checked": date.today().isoformat()}
    DATA.mkdir(exist_ok=True)
    SERPER_QUOTA_FILE.write_text(json.dumps(data, ensure_ascii=False),
                                 encoding="utf-8")
    log(f"[SERPER] Kredyty: zuzyte {_serper_used['n']} w tym runie, "
        f"lacznie {used} z {total} (pozostalo {data['left']})")


# ========================================================== extraction ======

PRODUCT_TYPES = {"product", "productgroup", "offer", "aggregateoffer"}


def _iter_jsonld_nodes(soup):
    """Iteruje po wezlach JSON-LD typu Product/ProductGroup/Offer, zbiera tez
    liste wszystkich napotkanych typow (diagnostyka)."""
    seen_types = set()
    nodes = []
    for tag in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(tag.string or "")
        except Exception:
            continue
        stack = data if isinstance(data, list) else [data]
        while stack:
            node = stack.pop()
            if not isinstance(node, dict):
                continue
            if "@graph" in node:
                stack += [n for n in node["@graph"] if isinstance(n, dict)]
            t = node.get("@type", "")
            types = t if isinstance(t, list) else [t]
            tl = {str(x).lower() for x in types if x}
            seen_types |= tl
            if tl & PRODUCT_TYPES:
                nodes.append(node)
    return nodes, seen_types


def _norm_availability(val):
    if not val:
        return ""
    # "In Stock", "out of stock", "https://schema.org/InStock" -> "instock"...
    return re.sub(r"[^a-z]", "", str(val).rsplit("/", 1)[-1].lower())


def _parse_price(val):
    if val is None:
        return None
    s = str(val).strip().replace("\xa0", "").replace(" ", "")
    s = re.sub(r"[^\d,.\-]", "", s)
    if "," in s and "." in s:
        s = s.replace(",", "") if s.rfind(".") > s.rfind(",") else s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        v = float(s)
        return v if v > 0 else None
    except ValueError:
        return None


def _ident_of(obj):
    parts = []
    for k in ("sku", "name", "mpn", "url", "@id"):
        v = obj.get(k)
        if isinstance(v, dict):
            v = v.get("name", "")
        if v:
            parts.append(str(v))
    return " ".join(parts).lower()


def _collect_offers(obj, ident, out):
    """Zbiera oferty z pola offers danego obiektu (obsluga AggregateOffer,
    zagniezdzonych offers[] i priceSpecification)."""
    offs = obj.get("offers")
    if not offs:
        return
    offs = list(offs) if isinstance(offs, list) else [offs]
    for off in offs:
        if not isinstance(off, dict):
            continue
        inner = off.get("offers")
        if isinstance(inner, list):
            offs += [o for o in inner if isinstance(o, dict)]
        price = _parse_price(off.get("price", off.get("lowPrice")))
        currency = str(off.get("priceCurrency", "")).upper().strip()
        if price is None or not currency:
            ps = off.get("priceSpecification")
            if isinstance(ps, list):
                ps = ps[0] if ps else None
            if isinstance(ps, dict):
                price = price if price is not None else _parse_price(ps.get("price"))
                currency = currency or str(ps.get("priceCurrency", "")).upper().strip()
        avail = _norm_availability(off.get("availability"))
        io = off.get("itemOffered")
        io_ident = _ident_of(io) if isinstance(io, dict) else ""
        if price and currency:
            out.append({"price": price, "currency": currency, "avail": avail,
                        "ident": (_ident_of(off) + " " + io_ident + " " + ident).strip()})


def _extract_next_data(soup):
    """Fallback: sklepy headless Next.js/Shopify (np. laufcycles.com) publikuja
    dane produktu w skrypcie __NEXT_DATA__ zamiast w JSON-LD. Czytamy tylko
    props.pageProps.product (glowny produkt strony), zeby nie zlapac cen
    produktow powiazanych/polecanych."""
    tag = soup.find("script", id="__NEXT_DATA__")
    if not tag:
        return []
    try:
        data = json.loads(tag.string or "")
    except Exception:
        return []
    prod = ((data.get("props") or {}).get("pageProps") or {}).get("product")
    if not isinstance(prod, dict):
        return []
    variants = prod.get("variants")
    if not isinstance(variants, list) or not variants:
        variants = [prod]
    out = []
    for v in variants:
        if not isinstance(v, dict):
            continue
        pr = v.get("price")
        if not isinstance(pr, dict):
            continue
        price = _parse_price(pr.get("amount"))
        currency = str(pr.get("currencyCode", "")).upper().strip()
        if not price or not currency:
            continue
        afs = v.get("availableForSale")
        avail = "" if afs is None else ("instock" if afs else "outofstock")
        ident = " ".join(str(v.get(k) or "") for k in
                         ("sku", "title", "size", "color")).strip().lower()
        out.append({"price": price, "currency": currency,
                    "avail": avail, "ident": ident})
    return out


def extract_offers(html):
    """Zwraca (lista_kandydatow, typy_jsonld). Kandydat: dict(price, currency,
    avail, ident) - ident sluzy dopasowaniu wariantu (np. rozmiaru)."""
    soup = BeautifulSoup(html, "html.parser")
    nodes, seen_types = _iter_jsonld_nodes(soup)
    out = []
    for node in nodes:
        ident = _ident_of(node)
        t = node.get("@type", "")
        types = {str(x).lower() for x in (t if isinstance(t, list) else [t])}
        if types & {"offer", "aggregateoffer"}:  # samodzielny wezel Offer
            price = _parse_price(node.get("price", node.get("lowPrice")))
            currency = str(node.get("priceCurrency", "")).upper().strip()
            avail = _norm_availability(node.get("availability"))
            if price and currency:
                out.append({"price": price, "currency": currency,
                            "avail": avail, "ident": ident})
        _collect_offers(node, ident, out)
        for v in (node.get("hasVariant") or []):        # ProductGroup
            if isinstance(v, dict):
                _collect_offers(v, _ident_of(v), out)

    if not out:  # fallback: meta OpenGraph / microdata
        mp = (soup.find("meta", attrs={"property": "product:price:amount"})
              or soup.find("meta", attrs={"property": "og:price:amount"})
              or soup.find("meta", attrs={"itemprop": "price"}))
        mc = (soup.find("meta", attrs={"property": "product:price:currency"})
              or soup.find("meta", attrs={"property": "og:price:currency"})
              or soup.find("meta", attrs={"itemprop": "priceCurrency"}))
        ma = (soup.find("meta", attrs={"property": "product:availability"})
              or soup.find("link", attrs={"itemprop": "availability"})
              or soup.find("meta", attrs={"itemprop": "availability"}))
        if mp and mc:
            price = _parse_price(mp.get("content"))
            currency = str(mc.get("content", "")).upper().strip()
            avail = _norm_availability(ma.get("content") or ma.get("href")) if ma else ""
            if price and currency:
                out.append({"price": price, "currency": currency,
                            "avail": avail, "ident": ""})

    if not out:  # fallback: headless Next.js/Shopify (__NEXT_DATA__)
        out = _extract_next_data(soup)
        if out:
            seen_types.add("next_data")
    return out, seen_types


# ========================================================== monitoring ======

MAX_CRAWL_PER_PAGE = 3      # ile linkow produktowych podjac z jednej strony
MAX_CRAWL_PER_PRODUCT = 6   # limit dodatkowych pobran per produkt / run


def _product_tokens(p):
    """Tokeny identyfikujace produkt (z id/nazwy/fraz): kody modelu z cyfra
    (np. 'mx190', 'fm2797') oraz slowa >=4 znaki (np. 'cutthroat', 'wide')."""
    text = " ".join([str(p.get("id") or ""), str(p.get("name") or "")]
                    + [str(q) for q in (p.get("queries") or [])])
    toks = set(re.findall(r"[a-z0-9]+", text.lower()))
    return {t for t in toks
            if (len(t) >= 3 and any(ch.isdigit() for ch in t)) or len(t) >= 4}


def find_product_links(html, base_url, tokens, excluded, worldwide, known):
    """Gdy strona nie ma ceny (kategoria, poradnik rozmiarow, artykul),
    szuka na niej linkow wygladajacych na karte TEGO produktu. Link musi
    zawierac w sciezce/tekscie token z cyfra (kod modelu) albo >=2 tokeny
    nazwy. Tylko ta sama domena, max MAX_CRAWL_PER_PAGE najlepszych."""
    soup = BeautifulSoup(html, "html.parser")
    base_dom = domain_of(base_url)
    scored = {}
    for a in soup.find_all("a", href=True):
        url = normalize_url(urljoin(base_url, a["href"])).rstrip("/")
        if not url.startswith("http") or domain_of(url) != base_dom:
            continue
        if url in known or url == base_url.rstrip("/"):
            continue
        if not url_allowed(url, excluded, worldwide):
            continue
        hay = (urlparse(url).path + " " + a.get_text(" ", strip=True)).lower()
        hits = {t for t in tokens if t in hay}
        strong = any(any(ch.isdigit() for ch in t) for t in hits)
        if not strong and len(hits) < 2:
            continue
        score = len(hits) + (2 if strong else 0)
        if score > scored.get(url, 0):
            scored[url] = score
    return sorted(scored, key=scored.get, reverse=True)[:MAX_CRAWL_PER_PAGE]


# --- Fallback przegladarkowy: headless Chromium (Playwright) ----------------
# Sklepy z ochrona anty-bot (Cloudflare itp.) odrzucaja requests nawet
# z pelnymi naglowkami, bo rozpoznaja fingerprint TLS biblioteki. Wtedy
# pobieramy strone prawdziwa przegladarka. Chromium startuje leniwie -
# tylko gdy cos faktycznie blokuje (w typowym runie ~kilka domen).
BLOCK_STATUSES = {403, 429, 503}
# Interstitiale anty-bota potrafia przyjsc z HTTP 200 (Akamai serwuje stub
# z <meta refresh ...bm-verify=...>, Cloudflare "Just a moment"). Rozpoznajemy
# je po sygnaturach w naglowku HTML, zeby nie ekstrahowac ceny ze strony-zapory
# i od razu przejsc na przegladarke. Sygnatury dobrane wasko (male ryzyko
# falszywych trafien na normalnej karcie produktu).
_CHALLENGE_RE = re.compile(
    r"just a moment|cf-chl|challenge-platform|cf-browser-verification|"
    r"checking your browser|verifying you are human|"        # Cloudflare
    r"bm-verify|/_sec/verify|"                               # Akamai Bot Manager
    r"captcha-delivery|datadome|px-captcha",                 # DataDome / PerimeterX
    re.I)


def _is_challenge(html):
    """True gdy HTML to strona-zapora anty-bota, a nie wlasciwa tresc."""
    return bool(html) and bool(_CHALLENGE_RE.search(html[:8000]))
_pw_state = {"pw": None, "browser": None, "ctx": None, "failed": False}
_blocked_doms = set()  # requests dostal blokade -> kolejne URL-e od razu przegladarka
_dead_doms = set()     # przegladarka tez polegla -> nie marnuj czasu w tym runie


def _browser_context():
    """Leniwie startuje wspolna przegladarke; None gdy Playwright niedostepny."""
    if _pw_state["failed"]:
        return None
    if _pw_state["ctx"] is None:
        try:
            from playwright.sync_api import sync_playwright
            pw = sync_playwright().start()
            browser = pw.chromium.launch(
                headless=True,
                args=["--disable-blink-features=AutomationControlled"])
            ctx = browser.new_context(
                user_agent=UA, locale="en-US",
                viewport={"width": 1366, "height": 900})
            ctx.add_init_script(
                "Object.defineProperty(navigator,'webdriver',"
                "{get:()=>undefined})")
            ctx.set_default_timeout(BROWSER_TIMEOUT * 1000)
            _pw_state.update(pw=pw, browser=browser, ctx=ctx)
            log("[BROWSER] Start headless Chromium (fallback na blokady botow)")
        except Exception as e:
            _pw_state["failed"] = True
            log(f"[BROWSER] Playwright niedostepny ({type(e).__name__}) - "
                f"zablokowanych sklepow nie obejde. "
                f"Instalacja: pip install playwright && playwright install chromium")
    return _pw_state["ctx"]


def fetch_via_browser(url):
    """Pobiera strone headless Chromium. Zwraca (html, final_url) lub (None, None)."""
    ctx = _browser_context()
    if ctx is None:
        return None, None
    page = ctx.new_page()
    try:
        # obrazki/fonty/media sa zbedne do ekstrakcji ceny - nie pobieraj
        page.route(re.compile(r"\.(png|jpe?g|webp|avif|gif|svg|woff2?|ttf|mp4)"
                              r"(\?|$)"),
                   lambda route: route.abort())
        resp = page.goto(url, wait_until="domcontentloaded")
        html = page.content()
        # challenge JS (np. Cloudflare "Just a moment...") - daj mu czas przejsc
        saw_challenge = False
        for _ in range(4):
            if not _is_challenge(html):
                break
            saw_challenge = True
            page.wait_for_timeout(4000)
            html = page.content()
        if _is_challenge(html):
            return None, None                       # challenge nieprzejscie
        if resp and resp.status >= 400 and not saw_challenge:
            return None, None                       # zwykla blokada bez challenge
        return html, normalize_url(page.url)
    except Exception as e:
        log(f"  ! przegladarka: blad ({type(e).__name__})")
        return None, None
    finally:
        page.close()


def close_browser():
    if _pw_state["ctx"] is not None:
        try:
            _pw_state["browser"].close()
            _pw_state["pw"].stop()
        except Exception:
            pass
        _pw_state.update(pw=None, browser=None, ctx=None)


def fetch_offers(products, sources, rates):
    """Zwraca (offers, blind): offers to lista dict(product_id, name, url,
    domain, price, currency, price_pln, availability); blind to mapa
    {product_id: [domeny]}, z ktorych zaden URL nie dal sie pobrac
    (blokada botow / blad HTTP) - kandydaci do sprawdzenia przez Google."""
    session = requests.Session()
    session.headers.update(BROWSER_HEADERS)
    offers = []
    rejected = set()  # (pid, dom) - sklep odrzucil pobranie

    for p in products:
        pid = p["id"]
        excluded = set(p.get("exclude_domains") or [])
        worldwide = bool(p.get("worldwide"))
        variant = str(p.get("variant") or "").strip().lower()
        strict = bool(p.get("variant_strict"))
        ean = str(p.get("ean") or "").strip()
        ean_strict = bool(p.get("ean_strict"))
        min_pln = float(p.get("min_pln") or 0)
        max_pln = float(p.get("max_pln") or 0)
        # require_tokens: kazdy wpis musi wystapic na stronie; "a|b" = a LUB b
        req_tokens = [[alt.strip().lower() for alt in str(t).split("|")
                       if alt.strip()]
                      for t in (p.get("require_tokens") or [])]
        allowed_cur = WORLDWIDE_CURRENCIES if worldwide else ALLOWED_CURRENCIES
        urls = dict(sources.get(pid, {}))
        for u in p.get("seed_urls") or []:
            u = normalize_url(u)
            urls.setdefault(u, {"domain": domain_of(u), "first_seen": "seed"})

        tokens = _product_tokens(p)
        queue = list(urls.items())
        crawl_budget = MAX_CRAWL_PER_PRODUCT
        log(f"[MONITOR] {pid}: sprawdzam {len(urls)} URL-i"
            + (f" (wariant: {variant})" if variant else ""))
        while queue:
            url, meta = queue.pop(0)
            dom = meta.get("domain") or domain_of(url)
            if not url_allowed(url, excluded, worldwide):
                continue
            if dom in _dead_doms:
                log(f"  - {dom}: pomijam (blokada potwierdzona w tym runie)")
                rejected.add((pid, dom))
                continue
            try:
                html = final_url = None
                status = 0
                if dom not in _blocked_doms:
                    r = session.get(url, timeout=TIMEOUT, allow_redirects=True)
                    status = r.status_code
                    if status == 200 and not _is_challenge(r.text):
                        html = r.text
                        final_url = normalize_url(r.url or url)  # po przekierowaniach
                    elif status == 200 or status in BLOCK_STATUSES:
                        _blocked_doms.add(dom)   # dalsze URL-e od razu przegladarka
                    else:
                        log(f"  - {dom}: HTTP {status}")
                        rejected.add((pid, dom))
                        continue
                if html is None:
                    # blokada botow -> proba prawdziwa przegladarka
                    html, final_url = fetch_via_browser(url)
                    if html is None:
                        _dead_doms.add(dom)
                        rejected.add((pid, dom))
                        why = f"HTTP {status}" if status else "blokada botow"
                        log(f"  - {dom}: {why} (przegladarka tez nie przeszla)")
                        continue
                    log(f"  + {dom}: blokada ominieta headless Chromium")
                cands, seen_types = extract_offers(html)
            except Exception as e:
                log(f"  - {dom}: blad pobierania ({type(e).__name__})")
                rejected.add((pid, dom))
                continue
            finally:
                time.sleep(SLEEP_BETWEEN_FETCHES)

            if not cands:
                # auto-crawl (1 poziom): moze to kategoria/artykul - poszukaj
                # na stronie linku do karty tego produktu
                links = []
                if crawl_budget > 0 and meta.get("via") != "crawl":
                    links = find_product_links(html, final_url, tokens,
                                               excluded, worldwide,
                                               set(urls))[:crawl_budget]
                if links:
                    crawl_budget -= len(links)
                    for link in links:
                        m = {"domain": domain_of(link),
                             "first_seen": date.today().isoformat(),
                             "via": "crawl"}
                        urls[link] = m
                        queue.append((link, m))
                    log(f"  ~ {dom}: brak ceny, ale znalazlem {len(links)} "
                        f"link(i) produktowe - sprawdzam")
                else:
                    types = ", ".join(sorted(seen_types)) or "brak JSON-LD"
                    log(f"  - {dom}: brak danych o cenie (typy na stronie: {types})")
                continue

            # weryfikacja EAN: kod na stronie = niemal pewne, ze to TEN produkt;
            # brak kodu to tylko ostrzezenie (chyba ze ean_strict)
            enote = ""
            if ean:
                if ean in html:
                    enote = " [EAN OK]"
                elif ean_strict:
                    log(f"  - {dom}: EAN {ean} nieobecny na stronie "
                        f"- pomijam (ean_strict)")
                    continue
                else:
                    enote = " [EAN niepotwierdzony]"

            # require_tokens: slowa musza wystapic w URL-u, tytule strony
            # lub identyfikatorze oferty (np. "carbon", "s14", "kit")
            if req_tokens:
                tm = re.search(r"<title[^>]*>(.*?)</title>", html, re.I | re.S)
                page_txt = (final_url + " " + (tm.group(1) if tm else "")).lower()
                matching = [c for c in cands
                            if all(any(alt in page_txt + " " + c["ident"]
                                       for alt in grp) for grp in req_tokens)]
                if not matching:
                    missing = [grp[0] for grp in req_tokens
                               if not any(alt in page_txt for alt in grp)]
                    log(f"  - {dom}: brak wymaganych slow "
                        f"({', '.join(missing) or 'w ofertach'}) - pomijam")
                    continue
                cands = matching

            vnote = ""
            if variant:
                matching = [c for c in cands if variant in c["ident"]]
                # strona publikuje warianty tylko, gdy jest >1 odrebny
                # identyfikator ofert; pojedyncza oferta z sku/nazwa produktu
                # to cena zbiorcza (np. WooCommerce bez rozbicia na rozmiary)
                has_variants = len({c["ident"] for c in cands if c["ident"]}) > 1
                if matching:
                    cands = matching
                    vnote = f" [wariant {variant} dopasowany]"
                elif has_variants:
                    log(f"  - {dom}: brak wariantu '{variant}' wsrod ofert - pomijam")
                    continue
                else:
                    # cena zbiorcza - ale niektore sklepy (np. Shopware) maja
                    # osobny URL per rozmiar: wariant widac w URL-u lub tytule
                    # (".../...-m-56-cm", "Rahmenkit 29 M / 56 cm")
                    tm = re.search(r"<title[^>]*>(.*?)</title>", html,
                                   re.I | re.S)
                    page_txt = (final_url + " "
                                + (tm.group(1) if tm else "")).lower()
                    if re.search(r"(?<![0-9a-z])" + re.escape(variant)
                                 + r"(?![0-9a-z])", page_txt):
                        vnote = f" [wariant {variant} w URL/tytule]"
                    elif strict:
                        log(f"  - {dom}: cena zbiorcza bez informacji o wariantach "
                            f"- pomijam (variant_strict)")
                        continue
                    else:
                        vnote = f" [wariant {variant} NIEZWERYFIKOWANY - cena zbiorcza]"

            in_stock = [c for c in cands if not c["avail"] or c["avail"] in IN_STOCK]
            if not in_stock:
                log(f"  - {dom}: niedostepny ({cands[0]['avail']})")
                continue
            valid_cur = [c for c in in_stock if c["currency"] in allowed_cur]
            if not valid_cur:
                log(f"  - {dom}: waluta {in_stock[0]['currency']} poza dozwolonym "
                    f"regionem - pomijam")
                continue
            usable = [c for c in valid_cur if c["currency"] in rates]
            if not usable:
                log(f"  - {dom}: brak kursu {valid_cur[0]['currency']}")
                continue
            # widelki cenowe (PLN): odrzuca czesci zamienne / zestawy / bledy
            if min_pln or max_pln:
                in_range = [c for c in usable
                            if (not min_pln
                                or c["price"] * rates[c["currency"]] >= min_pln)
                            and (not max_pln
                                 or c["price"] * rates[c["currency"]] <= max_pln)]
                if not in_range:
                    w = min(usable, key=lambda c: c["price"] * rates[c["currency"]])
                    log(f"  - {dom}: {round(w['price'] * rates[w['currency']], 2)} "
                        f"PLN poza widelkami "
                        f"[{min_pln or '-'}..{max_pln or '-'}] - pomijam")
                    continue
                usable = in_range
            best = min(usable, key=lambda c: c["price"] * rates[c["currency"]])
            price, currency = best["price"], best["currency"]
            price_pln = round(price * rates[currency], 2)
            offers.append({"product_id": pid, "name": p["name"], "url": final_url,
                           "domain": dom, "price": price, "currency": currency,
                           "price_pln": price_pln,
                           "availability": best["avail"] or "instock"})
            if meta.get("via") == "crawl":  # znaleziony crawlem i ma cene
                sources.setdefault(pid, {})[url] = meta  # -> do bazy na stale
            log(f"  + {dom}: {price} {currency} = {price_pln} PLN{vnote}{enote}")

    # blind spoty: domeny odrzucone, ktore nie daly ceny zadnym innym URL-em
    have = {(o["product_id"], o["domain"]) for o in offers}
    blind = {}
    for pid, dom in rejected - have:
        blind.setdefault(pid, []).append(dom)
    for pid in blind:
        blind[pid].sort()
        log(f"[MONITOR] {pid}: blind spoty do sprawdzenia przez Google: "
            f"{', '.join(blind[pid])}")
    return offers, blind


def drop_outliers(offers):
    """Lagodne auto-odciecie: przy >=3 ofertach produktu cena ponizej 25%
    mediany to niemal na pewno inny produkt (czesc zamienna, akcesorium).
    Prawdziwe promocje (-40%, -60%) przechodza bez problemu."""
    prices = {}
    for o in offers:
        prices.setdefault(o["product_id"], []).append(o["price_pln"])
    medians = {pid: statistics.median(v) for pid, v in prices.items()
               if len(v) >= 3}
    kept = []
    for o in offers:
        med = medians.get(o["product_id"])
        if med and o["price_pln"] < 0.25 * med:
            log(f"  ! {o['product_id']}: odrzucam outlier {o['price_pln']} PLN "
                f"({o['domain']}) - ponizej 25% mediany ({round(med, 2)} PLN)")
            continue
        kept.append(o)
    return kept


# ============================== blind spoty: ceny przez Google ==============
# Sklepy z twarda ochrona anty-bot (Akamai na bike24, Cloudflare na r2-bike)
# odrzucaja i requests, i headless Chromium. Ich ceny bierzemy z Google:
#  1. Google Shopping (SerpAPI) - feed Merchant Center wysylany przez sam
#     sklep; cena oficjalna, swieza (opoznienie godziny), 1 zapytanie
#     pokrywa wszystkie blind spoty produktu. Budzet: settings.shopping_budget
#     zapytan per run + rezerwa settings.serpapi_reserve na discovery.
#  2. Google Programmable Search (GOOGLE_API_KEY+GOOGLE_CX) - pagemap wynikow
#     zawiera cene ze structured data zebrana przez Googlebota (sklepy go
#     nie blokuja). 100 zapytan/dzien za darmo; cena moze byc sprzed kilku
#     dni (czestotliwosc crawla), stad to warstwa druga.

_GL_CURRENCY = {"pl": "PLN", "de": "EUR", "at": "EUR", "be": "EUR",
                "nl": "EUR", "fr": "EUR", "it": "EUR", "es": "EUR",
                "pt": "EUR", "ie": "EUR", "fi": "EUR", "sk": "EUR",
                "si": "EUR", "hr": "EUR", "lt": "EUR", "lv": "EUR",
                "ee": "EUR", "gr": "EUR", "lu": "EUR", "cz": "CZK",
                "ch": "CHF", "dk": "DKK", "se": "SEK", "no": "NOK",
                "hu": "HUF", "ro": "RON", "bg": "BGN",
                "gb": "GBP", "uk": "GBP"}


def _shopping_currency(price_str, gl):
    """Waluta oferty Shopping: symbol w cenie, fallback kraj zapytania."""
    s = (price_str or "").lower()
    if "zł" in s or "pln" in s:
        return "PLN"
    if "€" in s or "eur" in s:
        return "EUR"
    if "chf" in s:
        return "CHF"
    if "£" in s or "gbp" in s:
        return "GBP"
    if "kč" in s or "czk" in s:
        return "CZK"
    if "$" in s or "usd" in s:
        return "USD"
    return _GL_CURRENCY.get((gl or "").lower(), "")


def _serpapi_quota_left():
    try:
        return json.loads(QUOTA_FILE.read_text(encoding="utf-8")).get("left")
    except Exception:
        return None


def _google_query(p):
    """Najprecyzyjniejsze zapytanie o produkt: EAN > pierwsza fraza > nazwa."""
    if p.get("ean"):
        return str(p["ean"])
    qs = p.get("queries") or []
    return str(qs[0]) if qs else str(p["name"])


def _title_ok(p, text):
    """Filtry jakosci na tytule oferty z Google (odpowiednik require_tokens
    + wariant + tokeny nazwy z monitoringu). Zwraca (ok, notka_do_logu)."""
    text = text.lower()
    # tytul musi wygladac na TEN produkt: kod modelu z cyfra albo >=2 tokeny
    tokens = _product_tokens(p)
    hits = {t for t in tokens if t in text}
    strong = any(any(ch.isdigit() for ch in t) for t in hits)
    if not strong and len(hits) < 2:
        return False, "tytul nie pasuje do produktu"
    req = [[alt.strip().lower() for alt in str(t).split("|") if alt.strip()]
           for t in (p.get("require_tokens") or [])]
    if any(not any(alt in text for alt in grp) for grp in req):
        return False, "brak wymaganych slow"
    variant = str(p.get("variant") or "").strip().lower()
    if variant:
        if re.search(r"(?<![0-9a-z])" + re.escape(variant) + r"(?![0-9a-z])",
                     text):
            return True, f" [wariant {variant} w tytule]"
        if p.get("variant_strict"):
            return False, f"brak wariantu '{variant}' w tytule (variant_strict)"
        return True, f" [wariant {variant} NIEZWERYFIKOWANY - tytul feedu]"
    return True, ""


def _price_in_range(p, price_pln):
    min_pln = float(p.get("min_pln") or 0)
    max_pln = float(p.get("max_pln") or 0)
    return ((not min_pln or price_pln >= min_pln)
            and (not max_pln or price_pln <= max_pln))


def _to_float(s):
    """Cena z tekstu: radzi sobie z '1299.00', '1.299,00', '1 299,00'."""
    s = str(s).replace("\xa0", "").replace(" ", "")
    if "," in s and "." in s:
        s = (s.replace(".", "") if s.rfind(",") > s.rfind(".")
             else s.replace(",", ""))
    return float(s.replace(",", "."))


def _shopping_providers(settings):
    """Dostawcy Google Shopping w kolejnosci uzycia: Serper.dev (osobna,
    duza pula darmowych zapytan) przed SerpAPI (chronionym rezerwa na
    discovery). Nastepny dostawca wchodzi, gdy poprzedni nic nie zwrocil."""
    provs = []
    k = os.environ.get("SERPER_API_KEY", "").strip()
    if k:
        provs.append(("serper",
                      lambda q, gl: serper_shopping_items(q, k, gl)))
    k = os.environ.get("SERPAPI_KEY", "").strip()
    if k:
        reserve = int(settings.get("serpapi_reserve", 10))
        left = _serpapi_quota_left()
        if left is None or int(left) - reserve > 0:
            provs.append(("serpapi",
                          lambda q, gl: serpapi_shopping_items(q, k, gl)))
        else:
            log(f"[SHOPPING] SerpAPI na wyczerpaniu (pozostalo {left}, "
                f"rezerwa {reserve}) - pomijam ten silnik")
    return provs


def _shopping_fill(products_by_id, blind, rates, settings):
    """Warstwa 1: Google Shopping (Serper/SerpAPI).
    Zwraca (oferty, pokryte_pary)."""
    providers = _shopping_providers(settings)
    if not providers:
        return [], set()
    budget = int(settings.get("shopping_budget", 4))
    if budget <= 0:
        return [], set()
    eu = [str(c).lower() for c in (settings.get("serpapi_countries")
                                   or ["de", "pl"])]
    offers, covered = [], set()
    for pid, doms in blind.items():
        p = products_by_id[pid]
        allowed_cur = (WORLDWIDE_CURRENCIES if p.get("worldwide")
                       else ALLOWED_CURRENCIES)
        # kraj zapytania wg TLD sklepu (bike24.de -> gl=de, bikero.cz ->
        # gl=cz - kazdy kraj Europy); TLD neutralne (.com) -> eu[0]
        by_gl = {}
        for dom in doms:
            tld = dom.rsplit(".", 1)[-1]
            by_gl.setdefault(tld if tld in _GL_CURRENCY else eu[0],
                             set()).add(dom)
        query = _google_query(p)
        for gl, want in by_gl.items():
            items, used = [], None
            for pname, fn in providers:
                if budget <= 0:
                    log("[SHOPPING] Budzet zapytan na ten run wyczerpany")
                    return offers, covered
                budget -= 1
                items, used = fn(query, gl), pname
                if items:
                    break
            log(f"[SHOPPING] {pid} (gl={gl}, {used}): {len(items)} ofert "
                f"w feedzie, szukam: {', '.join(sorted(want))}")
            best = {}
            for it in items:
                if it.get("second_hand_condition"):
                    continue
                # dopasowanie do blind spotu: domena linku albo nazwa sklepu
                # z feedu ("BIKE24" ~ bike24.de, "r2-bike.com" ~ r2-bike.com)
                link = it.get("link") or ""
                dom = domain_of(link) if link.startswith("http") else ""
                src = re.sub(r"\W", "", str(it.get("source") or "").lower())
                hit = next((d for d in want if d == dom
                            or (len(src) >= 4
                                and (src in re.sub(r"\W", "", d)
                                     or re.sub(r"\W", "", d) in src))), None)
                if not hit:
                    continue
                ok, note = _title_ok(p, str(it.get("title") or ""))
                if not ok:
                    log(f"  - {hit}: {note} ('{it.get('title', '')[:60]}')")
                    continue
                # SerpAPI daje extracted_price; Serper tylko tekst ceny
                price = it.get("extracted_price")
                if not isinstance(price, (int, float)):
                    try:
                        price = _to_float(re.sub(r"[^\d.,]", "",
                                                 str(it.get("price") or "")))
                    except ValueError:
                        continue
                cur = _shopping_currency(str(it.get("price") or ""), gl)
                if price <= 0 or cur not in allowed_cur or cur not in rates:
                    continue
                price_pln = round(price * rates[cur], 2)
                if not _price_in_range(p, price_pln):
                    log(f"  - {hit}: {price_pln} PLN poza widelkami - pomijam")
                    continue
                # link bywa karta produktu Google - wtedy lepszy product_link
                url = (link if link and "google." not in dom
                       else it.get("product_link") or link or "")
                o = {"product_id": pid, "name": p["name"], "url": url,
                     "domain": hit, "price": price, "currency": cur,
                     "price_pln": price_pln, "availability": "instock",
                     "via": "shopping"}
                o["_note"] = note
                if hit not in best or price_pln < best[hit]["price_pln"]:
                    best[hit] = o
            for dom, o in best.items():
                note = o.pop("_note", "")
                covered.add((pid, dom))
                offers.append(o)
                log(f"  + {dom}: {o['price']} {o['currency']} = "
                    f"{o['price_pln']} PLN [Google Shopping]{note}")
            time.sleep(0.4)
    return offers, covered


def _cse_fill(products_by_id, blind, rates, settings):
    """Warstwa 2: Google Programmable Search - cena z indeksu Googlebota
    (pagemap offer/metatags). Dla par (produkt, domena) bez ceny z Shopping."""
    key = os.environ.get("GOOGLE_API_KEY", "").strip()
    cx = os.environ.get("GOOGLE_CX", "").strip()
    if not (key and cx):
        return []
    budget = int(settings.get("cse_budget", 15))
    offers = []
    for pid, doms in blind.items():
        p = products_by_id[pid]
        allowed_cur = (WORLDWIDE_CURRENCIES if p.get("worldwide")
                       else ALLOWED_CURRENCIES)
        query = _google_query(p)
        for dom in doms:
            if budget <= 0:
                log("[CSE] Budzet zapytan na ten run wyczerpany")
                return offers
            if "google" in _auth_dead:
                return offers
            budget -= 1
            try:
                r = requests.get("https://www.googleapis.com/customsearch/v1",
                                 params={"key": key, "cx": cx, "q": query,
                                         "num": 5, "siteSearch": dom,
                                         "siteSearchFilter": "i"},
                                 timeout=TIMEOUT)
                if r.status_code == 429:
                    log("[CSE] Limit dzienny wyczerpany")
                    return offers
                r.raise_for_status()
                items = r.json().get("items", [])
            except Exception as e:
                if _auth_fail("google", e, _HINT_GOOGLE):
                    return offers
                log(f"[CSE] {pid}/{dom}: blad zapytania ({type(e).__name__})")
                continue
            best = None
            for it in items:
                link = it.get("link") or ""
                # siteSearch filtruje po stronie Google; kontrola defensywna
                if not domain_of(link).endswith(dom):
                    continue
                ok, note = _title_ok(p, (it.get("title") or "") + " " + link)
                if not ok:
                    continue
                pm = it.get("pagemap") or {}
                cands = []
                for of in pm.get("offer") or []:
                    avail = re.sub(r"\W", "", str(of.get("availability")
                                                  or "").lower())
                    if "outofstock" in avail or "discontinued" in avail:
                        continue
                    try:
                        cands.append((_to_float(of.get("price", "")),
                                      str(of.get("pricecurrency")
                                          or "").upper()))
                    except ValueError:
                        continue
                mt = (pm.get("metatags") or [{}])[0]
                for pk, ck in (("product:price:amount",
                                "product:price:currency"),
                               ("og:price:amount", "og:price:currency")):
                    try:
                        cands.append((_to_float(mt.get(pk, "")),
                                      str(mt.get(ck) or "").upper()))
                    except ValueError:
                        continue
                for price, cur in cands:
                    if price <= 0 or cur not in allowed_cur or cur not in rates:
                        continue
                    price_pln = round(price * rates[cur], 2)
                    if not _price_in_range(p, price_pln):
                        continue
                    o = {"product_id": pid, "name": p["name"],
                         "url": it.get("link") or "", "domain": dom,
                         "price": price, "currency": cur,
                         "price_pln": price_pln, "availability": "instock",
                         "via": "google", "_note": note}
                    if best is None or price_pln < best["price_pln"]:
                        best = o
            if best:
                note = best.pop("_note", "")
                offers.append(best)
                log(f"  + {dom}: {best['price']} {best['currency']} = "
                    f"{best['price_pln']} PLN [indeks Google - cena moze byc "
                    f"sprzed kilku dni]{note}")
            else:
                log(f"[CSE] {pid}/{dom}: brak ceny w indeksie Google")
            time.sleep(0.5)
    return offers


def fill_blind_spots(products, blind, rates, settings):
    """Laczy warstwy: Shopping -> CSE. Zwraca oferty dla blind spotow."""
    if not blind:
        return []
    products_by_id = {p["id"]: p for p in products}
    offers, covered = _shopping_fill(products_by_id, blind, rates, settings)
    remaining = {pid: [d for d in doms if (pid, d) not in covered]
                 for pid, doms in blind.items()}
    remaining = {pid: doms for pid, doms in remaining.items() if doms}
    if remaining:
        offers += _cse_fill(products_by_id, remaining, rates, settings)
    return offers


# ============================================================ storage =======

HIST_COLS = ["date", "product_id", "product_name", "price_pln",
             "price_orig", "currency", "shop", "url", "offers_checked"]
OFFER_COLS = ["date", "product_id", "domain", "price", "currency",
              "price_pln", "availability", "url", "via"]


def append_history(offers, products, today):
    DATA.mkdir(exist_ok=True)
    # pelny audyt ofert; ponowny run tego samego dnia zastepuje wiersze z ta data
    old_offers = []
    if OFFERS_FILE.exists():
        with OFFERS_FILE.open(newline="", encoding="utf-8") as f:
            old_offers = [r for r in csv.DictReader(f) if r.get("date") != today]
    with OFFERS_FILE.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=OFFER_COLS)
        w.writeheader()
        w.writerows(old_offers)
        for o in offers:
            w.writerow({"date": today, "product_id": o["product_id"],
                        "domain": o["domain"], "price": o["price"],
                        "currency": o["currency"], "price_pln": o["price_pln"],
                        "availability": o["availability"], "url": o["url"],
                        "via": o.get("via", "")})

    # historia: najlepsza cena per produkt; nadpisz jesli dzis juz byl wpis
    rows = []
    if HISTORY_FILE.exists():
        with HISTORY_FILE.open(newline="", encoding="utf-8") as f:
            rows = [r for r in csv.DictReader(f)]
    rows = [r for r in rows if not (r["date"] == today)]

    by_product = {}
    for o in offers:
        cur = by_product.get(o["product_id"])
        if cur is None or o["price_pln"] < cur["price_pln"]:
            by_product[o["product_id"]] = o
    counts = {}
    for o in offers:
        counts[o["product_id"]] = counts.get(o["product_id"], 0) + 1

    for p in products:
        pid = p["id"]
        best = by_product.get(pid)
        if best:
            rows.append({"date": today, "product_id": pid,
                         "product_name": p["name"],
                         "price_pln": best["price_pln"],
                         "price_orig": best["price"],
                         "currency": best["currency"],
                         "shop": best["domain"], "url": best["url"],
                         "offers_checked": counts.get(pid, 0)})
            log(f"[HISTORIA] {pid}: najnizsza {best['price_pln']} PLN "
                f"({best['domain']})")
        else:
            log(f"[HISTORIA] {pid}: brak dostepnych ofert dzisiaj")

    rows.sort(key=lambda r: (r["product_id"], r["date"]))
    with HISTORY_FILE.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=HIST_COLS)
        w.writeheader()
        w.writerows(rows)
    return rows


# ========================================================= excel report =====

def safe_sheet_name(name):
    return re.sub(r"[\[\]:*?/\\]", "-", str(name))[:31]


def build_excel(rows, products):
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    OUTPUT.mkdir(exist_ok=True)
    wb = Workbook()

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", start_color="1F4E79")
    body_font = Font(name="Arial")
    money_fmt = "#,##0.00\\ \\z\\ł"

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    # --- Podsumowanie ---
    ws = wb.active
    ws.title = "Podsumowanie"
    ws.append(["Produkt", "Nazwa", "Ostatni pomiar", "Aktualna najniższa (PLN)",
               "Minimum historyczne (PLN)", "Data minimum", "Sklep (aktualny)", "Link"])
    style_header(ws, 8)

    per_product = {}
    for r in rows:
        per_product.setdefault(r["product_id"], []).append(r)

    for p in products:
        pid = p["id"]
        hist = per_product.get(pid, [])
        if not hist:
            ws.append([pid, p["name"], "-", "-", "-", "-", "-", "-"])
            continue
        last = hist[-1]
        m = min(hist, key=lambda r: float(r["price_pln"]))
        row = [pid, p["name"], last["date"], float(last["price_pln"]),
               float(m["price_pln"]), m["date"], last["shop"], last["url"]]
        ws.append(row)
        rr = ws.max_row
        ws.cell(rr, 4).number_format = money_fmt
        ws.cell(rr, 5).number_format = money_fmt
        link = ws.cell(rr, 8)
        link.hyperlink = last["url"]
        link.font = Font(name="Arial", color="0563C1", underline="single")

    widths = [14, 34, 14, 22, 24, 14, 26, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.font == Font():
                cell.font = body_font

    # --- Historia (pelna tabela) ---
    ws = wb.create_sheet("Historia")
    ws.append(["Data", "Produkt", "Nazwa", "Cena (PLN)", "Cena oryg.",
               "Waluta", "Sklep", "Link", "Ofert sprawdzonych"])
    style_header(ws, 9)
    for r in rows:
        ws.append([r["date"], r["product_id"], r["product_name"],
                   float(r["price_pln"]), float(r["price_orig"]), r["currency"],
                   r["shop"], r["url"], int(r.get("offers_checked") or 0)])
        ws.cell(ws.max_row, 4).number_format = money_fmt
    for i, w in enumerate([12, 14, 32, 14, 12, 8, 24, 48, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Arkusz per produkt z wykresem ---
    for p in products:
        pid = p["id"]
        hist = per_product.get(pid, [])
        ws = wb.create_sheet(safe_sheet_name(pid))
        ws.append(["Data", "Najniższa cena (PLN)", "Sklep", "Link"])
        style_header(ws, 4)
        for r in hist:
            ws.append([r["date"], float(r["price_pln"]), r["shop"], r["url"]])
            ws.cell(ws.max_row, 2).number_format = money_fmt
        for i, w in enumerate([12, 20, 26, 50], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        n = len(hist)
        if n >= 1:
            chart = LineChart()
            chart.title = f"{p['name']} - najniższa cena dzienna (PLN)"
            chart.style = 12
            chart.y_axis.title = "PLN"
            chart.x_axis.title = "Data"
            chart.height = 9
            chart.width = 22
            data = Reference(ws, min_col=2, min_row=1, max_row=1 + n)
            cats = Reference(ws, min_col=1, min_row=2, max_row=1 + n)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            s = chart.series[0]
            s.marker.symbol = "circle"
            s.marker.size = 6
            s.smooth = False
            ws.add_chart(chart, "F2")

    wb.save(EXCEL_FILE)
    log(f"[EXCEL] Zapisano {EXCEL_FILE}")


# =============================================================== main =======

def main():
    with PRODUCTS_FILE.open(encoding="utf-8") as f:
        cfg = yaml.safe_load(f) or {}
    products = cfg.get("products") or []
    settings = cfg.get("settings") or {}
    if not products:
        log("Brak produktow w products.yaml - koniec.")
        return 0
    # zduplikowane ID scalaja bazy URL-i i mieszaja oferty roznych produktow
    seen_ids, unique = set(), []
    for p in products:
        if p["id"] in seen_ids:
            log(f"[KONFIG] BLAD: zduplikowane id '{p['id']}' w products.yaml - "
                f"pomijam drugi wpis ('{p.get('name')}'). Nadaj mu unikalne id.")
            continue
        seen_ids.add(p["id"])
        unique.append(p)
    products = unique
    log(f"Produkty: {[p['id'] for p in products]}")

    DATA.mkdir(exist_ok=True)
    sources = {}
    if SOURCES_FILE.exists():
        sources = json.loads(SOURCES_FILE.read_text(encoding="utf-8"))
    sources = normalize_sources(sources)

    sources = run_discovery(products, sources, settings)
    SOURCES_FILE.write_text(json.dumps(sources, indent=2, ensure_ascii=False),
                            encoding="utf-8")
    save_serpapi_quota()

    rates = get_fx_rates()
    try:
        offers, blind = fetch_offers(products, sources, rates)
    finally:
        close_browser()
    # sklepy, ktore odrzucily pobranie -> ceny z Google (Shopping / indeks)
    offers += fill_blind_spots(products, blind, rates, settings)
    offers = drop_outliers(offers)
    save_serper_quota(settings)  # po discovery i blind spotach - pelne zuzycie

    # nierozwiazane blind spoty: sklep blokuje, a Google tez nie dal ceny.
    # Plik nadpisywany co run - sklep znika z listy, gdy tylko cena wroci
    # (dashboard pokazuje to w Diagnostyce jako liste domen do recznego
    # dopisania do wyszukiwarki PSE)
    have = {(o["product_id"], o["domain"]) for o in offers}
    unresolved = {pid: [d for d in doms if (pid, d) not in have]
                  for pid, doms in blind.items()}
    unresolved = {pid: doms for pid, doms in unresolved.items() if doms}
    BLIND_FILE.write_text(json.dumps(
        {"date": date.today().isoformat(), "spots": unresolved},
        indent=2, ensure_ascii=False), encoding="utf-8")
    for pid, doms in unresolved.items():
        log(f"[BLIND] {pid}: bez ceny (sklep blokuje, Google nie pokryl): "
            f"{', '.join(doms)}")
    # URL-e znalezione auto-crawlem w monitoringu -> utrwal w bazie
    SOURCES_FILE.write_text(json.dumps(sources, indent=2, ensure_ascii=False),
                            encoding="utf-8")
    today = date.today().isoformat()
    rows = append_history(offers, products, today)
    build_excel(rows, products)

    from dashboard import build_dashboard
    save_run_log()  # dashboard osadza log biezacego przebiegu (Diagnostyka)
    out = build_dashboard(rows, products, settings, today)
    log(f"[DASHBOARD] Zapisano {out} (+ docs/index.html dla GitHub Pages)")
    log("Gotowe.")
    save_run_log()
    return 0


if __name__ == "__main__":
    sys.exit(main())
